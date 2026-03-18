"""
XLSX Normalization.

Handles Microsoft Excel workbooks (.xlsx format).

Processing rules:

  1. FORMULA RESOLUTION
     openpyxl is opened with data_only=True so all cells return their
     last-computed value rather than the formula string.
     If a cell has no cached value (workbook was never saved after
     formula entry), the cell returns None — treated as empty string.

  2. SHEET ENUMERATION
     All worksheets are processed. Sheets where max_row < 2 or
     max_column < 2 are skipped as effectively empty.

  3. HEADER DETECTION
     Iterate from row 1 until a row where >= 60% of non-None cells
     contain non-numeric string values. That row is the header row.
     If no header row is found within the first 10 rows, row 1 is
     used as the header regardless.

  4. FINANCIAL MODEL DETECTION
     Sheet names are checked against a keyword list (case-insensitive):
       p&l, income, revenue, margin, ebitda, forecast, budget, model,
       bridge, waterfall, projection, summary
     Matching sheets are tagged table_type='financial_model'.

  5. LARGE SHEET SAMPLING
     Sheets with more than LARGE_SHEET_ROW_THRESHOLD rows (default 500)
     are sampled: first 50 rows, last 50 rows, middle 50 rows.
     The resulting ExtractedTable is tagged is_sampled=True.
     A warning TextBlock is injected noting the sampling.

  6. OUTPUT
     Each sheet produces one ExtractedTable.
     The sheet name is stored in the table's sheet_name field.
     A TextBlock header is produced per sheet summarising its content.

Output: NormalizedDocument with TextBlocks and ExtractedTables.
"""
from __future__ import annotations

import logging
import math
from typing import Any, Optional
from uuid import UUID, uuid4

import openpyxl
from openpyxl.utils import get_column_letter

from packages.core.schemas.document import (
    DocumentMetadata,
    ExtractedTable,
    NormalizedDocument,
    TextBlock,
)

logger = logging.getLogger(__name__)

# Thresholds
_HEADER_SEARCH_MAX_ROWS       = 10    # Search at most this many rows for header
_HEADER_STRING_RATIO          = 0.60  # >= this fraction of cells must be strings
_LARGE_SHEET_ROW_THRESHOLD    = 500   # Sheets with more rows than this are sampled
_SAMPLE_ROWS                  = 50    # Rows per sample window (first / mid / last)

# Financial model detection keywords (lowercase, checked against sheet name)
_FINANCIAL_KEYWORDS = {
    "p&l", "income", "revenue", "margin", "ebitda", "forecast",
    "budget", "model", "bridge", "waterfall", "projection", "summary",
    "pnl", "pl ", " pl", "profit", "loss",
}


def normalize_xlsx(
    content: bytes,
    document_id: UUID,
    engagement_id: UUID,
    filename: str,
    vdr_folder_path: Optional[str] = None,
) -> NormalizedDocument:
    """
    Normalize an XLSX file to a NormalizedDocument.

    Args:
        content:          Raw XLSX bytes.
        document_id:      UUID of the document record.
        engagement_id:    UUID of the engagement.
        filename:         Original filename (used in citations).
        vdr_folder_path:  Optional VDR folder path for metadata.

    Returns:
        NormalizedDocument with TextBlocks and ExtractedTables.
    """
    import io

    # data_only=True resolves formulas to cached computed values
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=False)

    text_blocks: list[TextBlock] = []
    tables: list[ExtractedTable] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Skip empty sheets
        if (ws.max_row is None or ws.max_row < 2 or
                ws.max_column is None or ws.max_column < 2):
            logger.debug("Skipping empty sheet: %s", sheet_name)
            continue

        table, sheet_blocks = _process_sheet(
            ws, sheet_name, document_id, filename
        )

        if table is not None:
            tables.append(table)
        text_blocks.extend(sheet_blocks)

    wb.close()

    return NormalizedDocument(
        document_id=document_id,
        engagement_id=engagement_id,
        text_blocks=text_blocks,
        tables=tables,
        metadata=DocumentMetadata(
            filename=filename,
            file_type="xlsx",
            page_count=None,
            vdr_folder_path=vdr_folder_path,
            section_path=[],
        ),
    )


# ─── Sheet processing ─────────────────────────────────────────────────────────

def _process_sheet(
    ws,
    sheet_name: str,
    document_id: UUID,
    filename: str,
) -> tuple[Optional[ExtractedTable], list[TextBlock]]:
    """
    Process one worksheet into an ExtractedTable and supporting TextBlocks.

    Returns (ExtractedTable or None, list[TextBlock]).
    """
    # Read all rows as lists of cell values
    all_rows = _read_sheet_rows(ws)
    if not all_rows:
        return None, []

    # Detect header row
    header_idx = _detect_header_row(all_rows)
    headers = [str(v) if v is not None else "" for v in all_rows[header_idx]]

    # Skip sheets where all header cells are empty
    if not any(h.strip() for h in headers):
        return None, []

    data_rows_raw = all_rows[header_idx + 1:]

    # Detect table type
    table_type = _detect_table_type(sheet_name)

    # Handle large sheets via sampling
    is_sampled = False
    sampled_warning: Optional[TextBlock] = None
    total_data_rows = len(data_rows_raw)

    if total_data_rows > _LARGE_SHEET_ROW_THRESHOLD:
        data_rows_raw = _sample_rows(data_rows_raw)
        is_sampled = True
        sampled_warning = TextBlock(
            block_id=uuid4(),
            document_id=document_id,
            page_number=None,
            heading_level=None,
            text=(
                f"[SAMPLED] Sheet '{sheet_name}' in {filename} has "
                f"{total_data_rows} data rows. "
                f"Showing first {_SAMPLE_ROWS}, middle {_SAMPLE_ROWS}, "
                f"and last {_SAMPLE_ROWS} rows only. "
                f"Claims citing this sheet should note data was sampled."
            ),
        )

    # Normalise all data rows to strings, enforce column count
    n_cols = len(headers)
    data_rows: list[list[str]] = []
    for raw_row in data_rows_raw:
        # Pad or trim to match header column count
        row = [str(v) if v is not None else "" for v in raw_row]
        if len(row) < n_cols:
            row.extend([""] * (n_cols - len(row)))
        elif len(row) > n_cols:
            row = row[:n_cols]
        data_rows.append(row)

    # Build the ExtractedTable
    table = ExtractedTable(
        table_id=uuid4(),
        document_id=document_id,
        page_number=None,
        sheet_name=sheet_name,
        headers=headers,
        rows=data_rows,
        table_type=table_type,
        is_sampled=is_sampled,
        is_malformed=False,
    )

    # Build a summary TextBlock for this sheet
    summary_text = (
        f"Sheet: {sheet_name} | "
        f"Columns: {', '.join(h for h in headers if h.strip())} | "
        f"Rows: {total_data_rows} data rows"
    )
    if table_type == "financial_model":
        summary_text += " [Financial Model]"

    summary_block = TextBlock(
        block_id=uuid4(),
        document_id=document_id,
        page_number=None,
        heading_level=3,   # Sheet name treated as H3
        text=summary_text,
    )

    text_blocks = [summary_block]
    if sampled_warning:
        text_blocks.append(sampled_warning)

    return table, text_blocks


def _read_sheet_rows(ws) -> list[list[Any]]:
    """
    Read all rows from a worksheet as a list of lists of cell values.
    Handles both regular and read-only worksheets.
    """
    rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


# ─── Header detection ─────────────────────────────────────────────────────────

def _detect_header_row(rows: list[list[Any]]) -> int:
    """
    Find the index of the header row.

    A header row has >= _HEADER_STRING_RATIO of its non-None cells
    containing non-numeric string values.

    Searches up to _HEADER_SEARCH_MAX_ROWS rows from the top.
    Falls back to row 0 if no header is detected.
    """
    search_limit = min(_HEADER_SEARCH_MAX_ROWS, len(rows))

    for idx in range(search_limit):
        row = rows[idx]
        non_none = [v for v in row if v is not None and str(v).strip()]
        if not non_none:
            continue
        string_count = sum(
            1 for v in non_none
            if isinstance(v, str) and not _is_numeric_string(v)
        )
        ratio = string_count / len(non_none)
        if ratio >= _HEADER_STRING_RATIO:
            return idx

    return 0   # Default: first row


def _is_numeric_string(s: str) -> bool:
    """Return True if the string represents a number (int, float, percentage)."""
    s = s.strip().rstrip("%").replace(",", "").replace("$", "")
    try:
        float(s)
        return True
    except ValueError:
        return False


# ─── Financial model detection ────────────────────────────────────────────────

def _detect_table_type(sheet_name: str) -> str:
    """
    Return 'financial_model' if sheet name contains any financial keyword,
    'generic' otherwise.
    """
    name_lower = sheet_name.lower()
    for keyword in _FINANCIAL_KEYWORDS:
        if keyword in name_lower:
            return "financial_model"
    return "generic"


# ─── Large sheet sampling ─────────────────────────────────────────────────────

def _sample_rows(rows: list[list[Any]]) -> list[list[Any]]:
    """
    Sample a large row list into first N, middle N, and last N rows.
    The three windows may overlap for very long sheets — deduplication
    is applied so each row appears at most once.
    """
    n = len(rows)
    s = _SAMPLE_ROWS

    first_end  = min(s, n)
    last_start = max(0, n - s)
    mid_start  = max(first_end, (n // 2) - (s // 2))
    mid_end    = min(last_start, mid_start + s)

    # Collect indices, preserving order, deduplicating
    seen: set[int] = set()
    indices: list[int] = []
    for i in (
        list(range(0, first_end)) +
        list(range(mid_start, mid_end)) +
        list(range(last_start, n))
    ):
        if i not in seen:
            seen.add(i)
            indices.append(i)

    return [rows[i] for i in sorted(indices)]
